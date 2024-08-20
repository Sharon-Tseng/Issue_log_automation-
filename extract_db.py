import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

issue_log_df = pd.read_excel(f"C:\\Users\\Sharon YY Tseng\\Downloads\\Model Office Trustee Issue Log 2024_LIVE (10).xlsx", 
                             sheet_name = "Model Office Issue Log", header = None)

trustee_status_df = pd.read_excel(f"C:\\Users\\Sharon YY Tseng\\Downloads\\Model Office Trustee Issue Log 2024_LIVE (10).xlsx",
                                  sheet_name = "Trustee_MO_status")

mod_log_df = issue_log_df.copy()
mod_log_df.columns = issue_log_df.iloc[3]

# Remove first 3 columns, and the last few columns 
mod_log_df.drop (columns = mod_log_df.columns[:1], axis = 1, inplace = True)
mod_log_df.drop(columns = mod_log_df.columns[16:], axis=1, inplace = True)
        
# function to extract individual trustee's LIVE issue log 
def extract_indi_log(input_df, trustee):
    indi_ls = []
    
    # Create a mask to handle NaN values
    mask = input_df["Issue ID (TR_No.)"].str.startswith(trustee) & ~input_df["Issue ID (TR_No.)"].isnull()
    
    if mask.any():
        indi_ls.append(input_df[mask])
    
    if indi_ls:
        indi_df = pd.concat(indi_ls)
    else:
        indi_df = pd.DataFrame(columns=input_df.columns)
    
    return indi_df

trustee_ongoing_ls = trustee_status_df[trustee_status_df["Status"]=="Ongoing"]

for i in trustee_ongoing_ls["Trustee"]:
    output_df = extract_indi_log(mod_log_df, i)
    
    # Export to Excel 
    output_df.to_excel(f"{i}_Model_Office_Trustee_Issue_Log.xlsx", index = False)
    print(f"Export Completed:{i}")




"""Specify colors: 
    light blue grey, green, 
    standard yellow (In Progress), standard blue (Closed), light blue (Proposed to close), light green (Ready to Re-run)"""
    
blue_grey_fill = PatternFill(start_color = "D9E1F2", end_color = "D9E1F2", fill_type = "solid")
green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type = "solid")

color_mapping = {
    "In Progress": "FFFF00", #standard yellow
    "Closed": "0320FF",  #standard blue
    "Propose to close": "00b0f0", #light blue
    "Ready to Re-run": "92D050",  # light green
    "Open": "ffc7ce"    # pink
}

# function to fill in color
def fill_color (sheet_name, min_row, min_col, max_col, color):
    for row in sheet_name.iter_rows(min_row = min_row, min_col=min_col, max_col = max_col):
        for cell in row:
            cell.fill = color
    
    
# import updated trustee issue logs
for i in trustee_ongoing_ls["Trustee"]:
    # Format the excel
    final_df = openpyxl.load_workbook(f"C:\\Users\\Sharon YY Tseng\\Desktop\\issue_log_automation\\"+str(i)+"_Model_Office_Trustee_Issue_Log.xlsx")
    final_sheet = final_df.active

    # Fill A column color
    fill_color(final_sheet,1,1,1,blue_grey_fill)
    fill_color(final_sheet, 1,9,9, blue_grey_fill)
    
    # Fill G~M column, E~G color
    fill_color(final_sheet,1,2,8, green_fill)
    fill_color(final_sheet,1,11,13,green_fill)
    
    # Fill in N column color
    for row in final_sheet.iter_rows(min_row=1, min_col=10, max_col=10):
        for cell in row:
            response = cell.value
            if response in color_mapping:
                color = color_mapping[response]
                fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.fill = fill

    final_df.save(f"{i}_Model_Office_Trustee_Issue_Log_formatted.xlsx")