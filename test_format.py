import openpyxl
import pandas as pd
from openpyxl import drawing 

issue_log_path = f"C:\\Users\\Sharon YY Tseng\\Desktop\\issue_log_automation\\Issue_log_automation-\\Model Office Trustee Issue Log 2024_LIVE.xlsx"
image_path = f"C:\\Users\\Sharon YY Tseng\\Desktop\\issue_log_automation\\Issue_log_automation-\\issue_severity_definition.png"

"""
Do not change the below

"""
# Load dest. workbook
log_wb = openpyxl.load_workbook(issue_log_path)
remove_ls = ["Sheet1", "Drop-down List", "Status update summary", "Issue Log Screenshot Reference ","Status", "Trustee_MO_status"]

for i in remove_ls:
    if i in log_wb.sheetnames:
        log_wb.remove(log_wb[i])
         
    
issue_ws = log_wb["Model Office Issue Log"]
log_wb.active = issue_ws

trustee_status_df = pd.read_excel(issue_log_path, sheet_name = "Trustee_MO_status")

# Identify selected trustee name
for i in trustee_status_df["Chosen_trustee:"].dropna():
    selected_trustee = i

# get max_row number 
total_rows = issue_ws.max_row

# Filter rows
for row_idx in range(total_rows, 4, -1): 
    cell_value = issue_ws.cell(row=row_idx, column=2).value
    if cell_value:
        issue_id_letters = ''.join(filter(str.isalpha, str(cell_value)))
        if issue_id_letters != selected_trustee:  
            issue_ws.delete_rows(row_idx, 1)

# Remove extra columns 
issue_ws = issue_ws.delete_cols(idx=18, amount=16)

# Add issue severity definition image
sev_def_ws = log_wb["Issue Severity Definition"]
log_wb.active = sev_def_ws

img = drawing.image.Image(image_path)
sev_def_ws.add_image(img, 'A1')

# Save workbook and close sessions     
log_wb.save(selected_trustee + "_issue_log.xlsx")       
log_wb.close()

